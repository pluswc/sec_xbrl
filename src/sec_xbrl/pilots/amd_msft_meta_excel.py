"""Export the AMD, MSFT, and META P2/P3 review as an Excel workbook.

The exporter is deliberately a presentation layer over :mod:`amd_msft_meta_p2`
and :mod:`amd_msft_meta_p3`: it neither creates mappings nor derives values.
"""

from __future__ import annotations

import argparse
import re
from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from sec_xbrl.pilots.amd_msft_meta_p1 import load_pilot_manifest
from sec_xbrl.pilots.amd_msft_meta_p2 import CompanyDossier, PilotP2Runner
from sec_xbrl.pilots.amd_msft_meta_p3 import (
    BacklogItem,
    P3PeerReview,
    PeerComparisonRow,
    build_peer_review,
)

SHEET_NAMES = (
    "Overview",
    "Company_Status",
    "Revenue_Breakdowns",
    "Disclosure_Status",
    "Peer_Comparison",
    "Source_Trace",
    "Backlog",
    "Revenue_Dashboard",
    "Revenue_Structure",
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_WARNING_FILL = PatternFill("solid", fgColor="FCE4D6")
_NOT_COMPARABLE_FILL = PatternFill("solid", fgColor="FFF2CC")
_LINK_FONT = Font(color="0563C1", underline="single")
_TAB_COLORS = {
    "Overview": "1F4E78", "Company_Status": "5B9BD5", "Revenue_Breakdowns": "70AD47",
    "Disclosure_Status": "70AD47", "Peer_Comparison": "ED7D31", "Source_Trace": "A5A5A5",
    "Backlog": "FFC000", "Revenue_Dashboard": "5B9BD5", "Revenue_Structure": "70AD47",
}
_REPORTED_VALUE_RE = re.compile(
    r"^(?P<number>[+-]?(?:\d+(?:\.\d*)?|\.\d+))(?:\s*(?:×|x)\s*10\^(?P<scale>[+-]?\d+))?$"
)
_NUMERIC_FORMAT = "#,##0.##########;[Red]-#,##0.##########"


@dataclass(frozen=True, slots=True)
class ReportedValue:
    """As-filed display and its Excel-calculable numeric representation."""

    numeric_value: Decimal
    as_filed_display: str
    scale: int | None


def parse_reported_value(value: str) -> ReportedValue:
    """Convert a P2/P3 reported display without discarding its lexical scale.

    A missing scale remains ``None`` while an explicit ``10^0`` remains ``0``.
    This distinction makes the raw display auditable even though both calculate
    to the same numeric value.
    """
    display = value.strip()
    match = _REPORTED_VALUE_RE.fullmatch(display)
    if not match:
        raise ValueError(f"unsupported reported value display: {value!r}")
    try:
        scale = int(match.group("scale")) if match.group("scale") is not None else None
        numeric_value = Decimal(match.group("number")) * (Decimal(10) ** (scale or 0))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"unsupported reported value display: {value!r}") from exc
    return ReportedValue(numeric_value=numeric_value, as_filed_display=display, scale=scale)


def export_workbook(
    *,
    dossiers: Iterable[CompanyDossier],
    review: P3PeerReview,
    filing_urls: Mapping[str, str],
    output: Path,
    relationship_evidence_status: str = "NOT_EVIDENCED",
) -> Path:
    """Write a reviewable workbook from already-produced P2/P3 evidence.

    ``filing_urls`` maps accession to the official SEC accession directory. It
    is used solely to make existing provenance clickable; no network request is
    performed.  ``output`` may be outside the repository's ignored artifacts
    directory for callers that need a different delivery location.
    """
    dossier_rows = tuple(sorted(dossiers, key=lambda row: row.ticker))
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in SHEET_NAMES:
        worksheet = workbook.create_sheet(name)
        worksheet.sheet_properties.tabColor = _TAB_COLORS[name]

    _overview(workbook["Overview"], review, filing_urls)
    _company_status(workbook["Company_Status"], dossier_rows, filing_urls)
    _revenue_breakdowns(workbook["Revenue_Breakdowns"], review, filing_urls)
    _disclosure_status(workbook["Disclosure_Status"], dossier_rows, filing_urls)
    _peer_comparison(workbook["Peer_Comparison"], review, filing_urls)
    _source_trace(workbook["Source_Trace"], review, filing_urls)
    _backlog(workbook["Backlog"], review)
    _revenue_dashboard(workbook["Revenue_Dashboard"], review)
    _revenue_structure(workbook["Revenue_Structure"], review, relationship_evidence_status)

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output


def export_pilot_workbook(*, manifest: Path, cache_root: Path, output: Path) -> Path:
    """Run the cached-only P2/P3 pilot and export its result to ``output``."""
    filings = tuple(load_pilot_manifest(manifest))
    dossiers = PilotP2Runner(cache_root=cache_root).run(filings)
    return export_workbook(
        dossiers=dossiers,
        review=build_peer_review(dossiers),
        filing_urls={row.filing.accession: row.filing_url for row in filings},
        output=output,
    )


def export_documented_review_workbook(
    *, manifest: Path, p2_summary: Path, p3_summary: Path, output: Path
) -> Path:
    """Export the committed P2/P3 review summaries without requiring raw cache.

    This is useful for distributing the already-reviewed pilot result. The
    cached-only runner remains the preferred command when regenerating P2/P3.
    """
    filings = tuple(load_pilot_manifest(manifest))
    dossiers = _summary_dossiers(filings, p2_summary.read_text(encoding="utf-8"))
    review = _summary_review(p3_summary.read_text(encoding="utf-8"))
    return export_workbook(
        dossiers=dossiers,
        review=review,
        filing_urls={row.filing.accession: row.filing_url for row in filings},
        output=output,
        relationship_evidence_status="DOCUMENTED_SUMMARY_ONLY",
    )


def _summary_dossiers(filings: Iterable[object], p2_summary: str) -> tuple[CompanyDossier, ...]:
    states = _summary_disclosure_states(p2_summary)
    by_ticker: dict[str, list[object]] = {}
    for pilot in filings:
        by_ticker.setdefault(pilot.ticker, []).append(pilot)
    result: list[CompanyDossier] = []
    for ticker, rows in sorted(by_ticker.items()):
        annual = next(row for row in rows if row.selection_role == "ANNUAL_BASELINE")
        update = next(row for row in rows if row.selection_role == "CURRENT_UPDATE")
        result.append(CompanyDossier(
            company=annual.company,
            ticker=ticker,
            annual_accession=annual.filing.accession,
            update_accession=update.filing.accession,
            annual_revenue=(), update_revenue=(), annual_breakdowns=(), update_breakdowns=(),
            disclosure_states=states.get(ticker, ()), statement_qa=(), relationship_qa=(),
            warnings=("Rendered from the committed P2/P3 review summaries; no new fact extraction was run.",),
        ))
    return tuple(result)


def _summary_disclosure_states(p2_summary: str) -> dict[str, tuple[tuple[str, str], ...]]:
    result: dict[str, tuple[tuple[str, str], ...]] = {}
    for ticker, reported, absent in re.findall(r"^\| (AMD|MSFT|META) \| (.*?) \| (.*?) \|$", p2_summary, re.MULTILINE):
        values = [(topic.strip(), "REPORTED_UNCHANGED") for topic in reported.split(",")]
        values += [(topic.strip(), "NOT_REPORTED_THIS_QUARTER") for topic in absent.split(",")]
        result[ticker] = tuple(values)
    return result


def _summary_review(p3_summary: str) -> P3PeerReview:
    comparisons: list[PeerComparisonRow] = []
    sections = re.findall(r"^### (AMD|MSFT|META) — `([^`]+)`\n\n(.*?)(?=^### |^## )", p3_summary, re.MULTILINE | re.DOTALL)
    for ticker, metric, section in sections:
        value, period_class, source_period, dimensions = re.search(
            r"Reported value: `(.*?)` \(`(.*?)`; (.*?); dimensions: (.*?)\)\.", section
        ).groups()
        raw_id = re.search(r"Raw ID: `(.*?)`", section).group(1)
        canonical_id = re.search(r"Company canonical ID: `(.*?)`", section).group(1)
        analytical_id, relation, confidence, version = re.search(
            r"Analytical ID: `(.*?)`; relation: `(.*?)`; confidence: `(.*?)`; version: `(.*?)`", section
        ).groups()
        evidence = re.search(r"Mapping evidence: (.*)", section).group(1)
        accession, document, locator, qname, unit = re.search(
            r"Source: accession `(.*?)`; document `(.*?)`; locator `(.*?)`; QName `(.*?)`; unit `(.*?)`", section
        ).groups()
        warning = re.search(r"Scope warning: (.*)", section).group(1)
        comparisons.append(PeerComparisonRow(
            ticker=ticker, metric=metric, value=value, period_class=period_class, source_period=source_period,
            source_raw_id=raw_id, company_canonical_id=canonical_id,
            analytical_id=None if analytical_id == "none" else analytical_id, mapping_relation=relation,
            mapping_confidence=float(confidence), mapping_version=version, mapping_evidence=evidence,
            accession=accession, source_document=document, source_locator=locator, qname=qname, unit=unit,
            dimensions=() if dimensions == "none" else tuple(dimensions.split("; ")), scope_warning=warning,
        ))
    warnings = tuple(re.findall(r"^- (.*)$", re.search(
        r"^## Panel-wide scope warnings\n\n(.*?)(?=^## )", p3_summary, re.MULTILINE | re.DOTALL
    ).group(1), re.MULTILINE))
    backlog: list[BacklogItem] = []
    for priority, item_id, section in re.findall(
        r"^### (.*?) — `(P3-[^`]+)`\n\n(.*?)(?=^### |\Z)", p3_summary, re.MULTILINE | re.DOTALL
    ):
        owner, lane = re.search(r"Owner lane: (.*?) \((.*?)\)\.", section).groups()
        gap = re.search(r"Evidence gap: (.*)", section).group(1)
        impact = re.search(r"Impact: (.*)", section).group(1)
        decision = re.search(r"Decision needed: (.*)", section).group(1)
        backlog.append(BacklogItem(priority, item_id, lane, owner, gap, impact, decision))
    return P3PeerReview(tuple(comparisons), tuple(backlog), warnings)


def _overview(sheet: object, review: P3PeerReview, filing_urls: Mapping[str, str]) -> None:
    rows = [
        ("AMD · MSFT · META pilot", "P2/P3 provenance-rich peer review"),
        ("Purpose", "Read reported facts together without asserting unreviewed comparability."),
        ("Period rule", "FY, QTD_3M, and YTD_9M are separate classes; do not aggregate or compare across classes."),
        ("Mapping rule", "UNMAPPED canonical IDs and UNRESOLVED relations are visible, not inferred."),
        ("Source boundary", "Six P1-validated cached filings; official SEC links are preserved per accession."),
        ("Available filings", ", ".join(sorted(filing_urls))),
    ]
    _write_key_values(sheet, rows)
    start = sheet.max_row + 2
    _write_table(sheet, start, ["Panel-wide scope warning"], [(warning,) for warning in review.warnings], "OverviewWarnings")
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 120


def _company_status(sheet: object, dossiers: Iterable[CompanyDossier], filing_urls: Mapping[str, str]) -> None:
    headers = ["Ticker", "Company", "Annual accession", "Annual SEC filing", "Current accession", "Current SEC filing", "P2 warning"]
    rows = []
    for dossier in dossiers:
        rows.append((dossier.ticker, dossier.company, dossier.annual_accession, "Open SEC filing", dossier.update_accession,
                     "Open SEC filing", " | ".join(dossier.warnings)))
    _write_table(sheet, 1, headers, rows, "CompanyStatus")
    for row_number, dossier in enumerate(dossiers, start=2):
        _set_link(sheet.cell(row_number, 4), filing_urls.get(dossier.annual_accession))
        _set_link(sheet.cell(row_number, 6), filing_urls.get(dossier.update_accession))


def _revenue_breakdowns(sheet: object, review: P3PeerReview, filing_urls: Mapping[str, str]) -> None:
    headers = [
        "Ticker", "Metric", "Numeric value", "As-filed display", "Unit", "Scale", "Period class",
        "Source period", "Dimensions", "Relation", "Confidence", "Scope warning", "SEC filing",
    ]
    comparisons = [row for row in review.comparisons if row.dimensions]
    rows = [
        (row.ticker, row.metric, *_reported_cells(row.value, row.unit), row.period_class, row.source_period,
         _dimensions(row.dimensions), row.mapping_relation, row.mapping_confidence, row.scope_warning, "Open SEC filing")
        for row in comparisons
    ]
    _write_table(sheet, 1, headers, rows, "RevenueBreakdowns")
    for row_number, row in enumerate(comparisons, start=2):
        _set_link(sheet.cell(row_number, 13), _document_url(filing_urls, row.accession, row.source_document))
    _format_numeric_column(sheet, "C")
    _highlight_mapping_relations(sheet, "J")


def _disclosure_status(sheet: object, dossiers: Iterable[CompanyDossier], filing_urls: Mapping[str, str]) -> None:
    headers = ["Ticker", "Disclosure topic", "P2 state", "Annual accession", "Annual SEC filing", "Current accession", "Current SEC filing"]
    rows = []
    link_rows: list[tuple[int, CompanyDossier]] = []
    for dossier in dossiers:
        for topic, state in dossier.disclosure_states:
            rows.append((dossier.ticker, topic, state, dossier.annual_accession, "Open SEC filing", dossier.update_accession, "Open SEC filing"))
            link_rows.append((len(rows) + 1, dossier))
    _write_table(sheet, 1, headers, rows, "DisclosureStatus")
    for row_number, dossier in link_rows:
        _set_link(sheet.cell(row_number, 5), filing_urls.get(dossier.annual_accession))
        _set_link(sheet.cell(row_number, 7), filing_urls.get(dossier.update_accession))
    _highlight_exact(sheet, "C", "NOT_REPORTED_THIS_QUARTER", _WARNING_FILL)


def _peer_comparison(sheet: object, review: P3PeerReview, filing_urls: Mapping[str, str]) -> None:
    headers = [
        "Ticker", "Metric", "Numeric value", "As-filed display", "Unit", "Scale", "Period class",
        "Source period", "Relation", "Confidence", "Mapping version", "Scope warning", "SEC filing",
    ]
    rows = [
        (row.ticker, row.metric, *_reported_cells(row.value, row.unit), row.period_class, row.source_period,
         row.mapping_relation, row.mapping_confidence, row.mapping_version, row.scope_warning, "Open SEC filing")
        for row in review.comparisons
    ]
    _write_table(sheet, 1, headers, rows, "PeerComparison")
    for row_number, row in enumerate(review.comparisons, start=2):
        _set_link(sheet.cell(row_number, 13), _document_url(filing_urls, row.accession, row.source_document))
    _format_numeric_column(sheet, "C")
    _highlight_mapping_relations(sheet, "I")


def _source_trace(sheet: object, review: P3PeerReview, filing_urls: Mapping[str, str]) -> None:
    headers = [
        "Ticker", "Metric", "Numeric value", "As-filed display", "Unit", "Scale", "Period class", "Source period",
        "Dimensions", "Raw fact ID", "Company canonical ID", "Analytical ID", "Mapping relation", "Mapping confidence",
        "Mapping version", "Mapping evidence", "Accession", "Source document", "Source locator", "QName", "SEC filing",
    ]
    rows = [
        (row.ticker, row.metric, *_reported_cells(row.value, row.unit), row.period_class, row.source_period,
         _dimensions(row.dimensions), row.source_raw_id, row.company_canonical_id, row.analytical_id or "none",
         row.mapping_relation, row.mapping_confidence, row.mapping_version, row.mapping_evidence, row.accession,
         row.source_document or "", row.source_locator or "", row.qname or "", "Open SEC filing")
        for row in review.comparisons
    ]
    _write_table(sheet, 1, headers, rows, "SourceTrace")
    for row_number, row in enumerate(review.comparisons, start=2):
        _set_link(sheet.cell(row_number, 21), _document_url(filing_urls, row.accession, row.source_document))
    _format_numeric_column(sheet, "C")
    _highlight_mapping_relations(sheet, "M")


def _backlog(sheet: object, review: P3PeerReview) -> None:
    headers = ["Priority", "Item ID", "Lane", "Owner lane", "Evidence gap", "Impact", "Decision needed"]
    rows = [(item.priority, item.item_id, item.lane, item.owner_lane, item.evidence_gap, item.impact, item.decision_needed) for item in review.backlog]
    _write_table(sheet, 1, headers, rows, "Backlog")
    _highlight_exact(sheet, "A", "P0 — correctness blocker", _WARNING_FILL)


def _revenue_dashboard(sheet: object, review: P3PeerReview) -> None:
    """Show only same-class QTD total revenue; it is not a peer ranking."""
    sheet["A1"] = "Revenue Dashboard — current QTD_3M reported total revenue"
    sheet["A2"] = "WARNING: side-by-side display only; all current total-revenue relations are UNRESOLVED, not EQUIVALENT."
    sheet["A3"] = "FY and YTD_9M are intentionally excluded from this chart and cards. No composition aggregate is calculated."
    sheet.merge_cells("A1:F1")
    sheet.merge_cells("A2:F2")
    sheet.merge_cells("A3:F3")
    for cell in (sheet["A1"], sheet["A2"], sheet["A3"]):
        cell.alignment = Alignment(wrap_text=True)
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"].fill = _WARNING_FILL
    rows = [row for row in review.comparisons if row.metric == "REPORTED_TOTAL_REVENUE" and row.period_class == "QTD_3M"]
    headers = ["Ticker", "Numeric value", "As-filed display", "Unit", "Period class", "Relation"]
    values = [(row.ticker, *_reported_cells(row.value, row.unit)[:3], row.period_class, row.mapping_relation) for row in rows]
    _write_table(sheet, 5, headers, values, "RevenueDashboard")
    _format_numeric_column(sheet, "B")
    _highlight_mapping_relations(sheet, "F")
    if rows:
        chart = BarChart()
        chart.title = "Current QTD_3M reported total revenue (not a ranking)"
        chart.y_axis.title = "Reported numeric value (USD)"
        chart.add_data(Reference(sheet, min_col=2, min_row=5, max_row=5 + len(rows)), titles_from_data=True)
        chart.set_categories(Reference(sheet, min_col=1, min_row=6, max_row=5 + len(rows)))
        chart.height = 7
        chart.width = 15
        sheet.add_chart(chart, "H5")


def _revenue_structure(sheet: object, review: P3PeerReview, relationship_status: str) -> None:
    """Render reported revenue rows without inferring a member hierarchy."""
    sheet["A1"] = "Revenue Structure — reported rows, not a composition aggregate"
    sheet["A2"] = "Display depth is a UI indentation only. Scope depth is the number of axis/member dimensions; it is not a parent/member or total relationship."
    sheet["A3"] = "PRE is presentation/context only and never expands inference. DEF requires allowed arcs and explicit targetRole; absent row evidence remains NOT_EVIDENCED."
    for row in (1, 2, 3):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True)
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A2"].fill = _WARNING_FILL
    headers = [
        "Ticker", "Period class", "Display depth", "Scope depth", "As-filed label", "Metric", "Numeric value",
        "As-filed display", "Unit", "Scale", "Dimensions", "Source locator", "Presentation evidence",
        "Presentation role", "Presentation relationship", "Definition evidence", "Definition role", "Definition relationship",
        "Relationship provenance",
    ]
    selected = [row for row in review.comparisons if row.metric == "REPORTED_TOTAL_REVENUE" or row.dimensions]
    selected.sort(key=lambda row: (row.ticker, row.period_class, row.metric != "REPORTED_TOTAL_REVENUE", row.metric))
    values = []
    for row in selected:
        depth = 0 if row.metric == "REPORTED_TOTAL_REVENUE" else 1
        evidence = "NOT_EVIDENCED" if relationship_status != "DOCUMENTED_SUMMARY_ONLY" else "NOT_AVAILABLE"
        values.append((
            row.ticker, row.period_class, depth, len(row.dimensions), "NOT_AVAILABLE", row.metric,
            *_reported_cells(row.value, row.unit), _dimensions(row.dimensions), row.source_locator or "",
            evidence, "NOT_AVAILABLE", "NOT_AVAILABLE", evidence, "NOT_AVAILABLE", "NOT_AVAILABLE", relationship_status,
        ))
    _write_table(sheet, 5, headers, values, "RevenueStructure")
    _format_numeric_column(sheet, "G")
    for row in range(6, sheet.max_row + 1):
        if sheet.cell(row, 3).value == 1:
            sheet.cell(row, 6).alignment = Alignment(indent=1, vertical="top", wrap_text=True)


def _write_key_values(sheet: object, rows: Iterable[tuple[str, str]]) -> None:
    for row in rows:
        sheet.append(row)
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _write_table(sheet: object, start: int, headers: list[str], rows: Iterable[tuple[object, ...]], table_name: str) -> None:
    for offset in range(start - 1):
        if offset >= sheet.max_row:
            sheet.append(())
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    end = sheet.max_row
    if end == start:
        sheet.append(tuple("" for _ in headers))
        end = sheet.max_row
    for cell in sheet[start]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ref = f"A{start}:{get_column_letter(len(headers))}{end}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)
    sheet.freeze_panes = f"A{start + 1}"
    sheet.auto_filter.ref = ref
    for row in sheet.iter_rows(min_row=start + 1, max_row=end):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column in range(1, len(headers) + 1):
        values = [str(sheet.cell(row, column).value or "") for row in range(start, end + 1)]
        width = min(max(max((len(value) for value in values), default=10) + 2, 12), 55)
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.sheet_view.showGridLines = False


def _set_link(cell: object, url: str | None) -> None:
    if url:
        cell.hyperlink = url
        cell.font = _LINK_FONT


def _document_url(filing_urls: Mapping[str, str], accession: str, document: str | None) -> str | None:
    filing_url = filing_urls.get(accession)
    if not filing_url:
        return None
    return filing_url.rstrip("/") + "/" + document if document else filing_url


def _dimensions(dimensions: tuple[str, ...]) -> str:
    return "; ".join(dimensions) if dimensions else "none"


def _reported_cells(value: str, unit: str | None) -> tuple[Decimal, str, str, int | None]:
    reported = parse_reported_value(value)
    return reported.numeric_value, reported.as_filed_display, unit or "", reported.scale


def _format_numeric_column(sheet: object, column: str) -> None:
    for row in range(2, sheet.max_row + 1):
        sheet[f"{column}{row}"].number_format = _NUMERIC_FORMAT


def _highlight_mapping_relations(sheet: object, column: str) -> None:
    _highlight_exact(sheet, column, "UNRESOLVED", _WARNING_FILL)
    _highlight_exact(sheet, column, "NOT_COMPARABLE", _NOT_COMPARABLE_FILL)


def _highlight_exact(sheet: object, column: str, value: str, fill: PatternFill) -> None:
    sheet.conditional_formatting.add(
        f"{column}2:{column}{sheet.max_row}",
        CellIsRule(operator="equal", formula=[f'"{value}"'], fill=fill),
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--manifest", type=Path, required=True, help="P0 pilot filing manifest")
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--cache-root", type=Path, help="Existing validated P1 cache (no download)")
    source.add_argument("--p3-summary", type=Path, help="Committed P3 review summary for cache-free export")
    parser.add_argument("--p2-summary", type=Path, help="Committed P2 review summary (required with --p3-summary)")
    parser.add_argument("--output", type=Path, default=Path("artifacts/AMD_MSFT_META_pilot.xlsx"))
    args = parser.parse_args(argv)
    if args.cache_root:
        export_pilot_workbook(manifest=args.manifest, cache_root=args.cache_root, output=args.output)
    else:
        if not args.p2_summary:
            parser.error("--p2-summary is required with --p3-summary")
        export_documented_review_workbook(
            manifest=args.manifest, p2_summary=args.p2_summary, p3_summary=args.p3_summary, output=args.output
        )
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(main())

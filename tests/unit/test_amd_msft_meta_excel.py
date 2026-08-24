from pathlib import Path

from openpyxl import load_workbook

from sec_xbrl.pilots.amd_msft_meta_excel import (
    SHEET_NAMES,
    export_documented_review_workbook,
    export_workbook,
)
from sec_xbrl.pilots.amd_msft_meta_p2 import CompanyDossier, DossierEvidence
from sec_xbrl.pilots.amd_msft_meta_p3 import build_peer_review


def _evidence(
    *, ticker: str, value: str, period_class: str, dimensions: tuple[str, ...] = ()
) -> DossierEvidence:
    return DossierEvidence(
        label="Revenue",
        value=value,
        period_class=period_class,
        period="2026-01-01 to 2026-03-31",
        dimensions=dimensions,
        accession=f"0000000000-26-00000{len(ticker)}",
        document=f"{ticker.lower()}.htm",
        locator="inline-id:f-1",
        qname="us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax",
        unit="iso4217:USD",
    )


def _dossier(ticker: str, evidence: DossierEvidence, breakdown: DossierEvidence | None = None) -> CompanyDossier:
    return CompanyDossier(
        company=f"{ticker} Example Corp.",
        ticker=ticker,
        annual_accession="0000000000-25-000001",
        update_accession=evidence.accession,
        annual_revenue=(),
        update_revenue=(evidence,),
        annual_breakdowns=(),
        update_breakdowns=(() if breakdown is None else (breakdown,)),
        disclosure_states=(("REVENUE_RECOGNITION", "NOT_REPORTED_THIS_QUARTER"),),
        statement_qa=(),
        relationship_qa=(),
        warnings=("Reported facts only.",),
    )


def test_exported_workbook_is_filterable_and_preserves_p2_p3_provenance(tmp_path: Path) -> None:
    amd = _dossier(
        "AMD",
        _evidence(ticker="AMD", value="10253 × 10^6", period_class="QTD_3M"),
        _evidence(
            ticker="AMD",
            value="5775 × 10^6",
            period_class="QTD_3M",
            dimensions=("us-gaap:StatementBusinessSegmentsAxis=amd:DataCenterMember",),
        ),
    )
    msft = _dossier("MSFT", _evidence(ticker="MSFT", value="241832 × 10^6", period_class="YTD_9M"))
    output = tmp_path / "AMD_MSFT_META_pilot.xlsx"
    export_workbook(
        dossiers=(amd, msft),
        review=build_peer_review((amd, msft)),
        filing_urls={
            amd.update_accession: "https://www.sec.gov/Archives/edgar/data/1/amd/",
            msft.update_accession: "https://www.sec.gov/Archives/edgar/data/2/msft/",
            amd.annual_accession: "https://www.sec.gov/Archives/edgar/data/1/annual/",
        },
        output=output,
    )

    workbook = load_workbook(output)
    assert tuple(workbook.sheetnames) == SHEET_NAMES
    assert all(workbook[name].freeze_panes for name in SHEET_NAMES)
    assert all(workbook[name].tables for name in SHEET_NAMES)

    trace = workbook["Source_Trace"]
    trace_headers = [cell.value for cell in trace[1]]
    for required in (
        "Raw fact ID", "Company canonical ID", "Mapping relation", "Mapping confidence",
        "Mapping version", "Source period", "Dimensions", "QName", "SEC filing",
    ):
        assert required in trace_headers
    assert "QTD_3M" in [cell.value for cell in trace[2]]
    assert "YTD_9M" in [cell.value for cell in trace[4]]
    assert trace.cell(2, trace_headers.index("SEC filing") + 1).hyperlink.target.endswith("/amd/amd.htm")

    peer = workbook["Peer_Comparison"]
    peer_headers = [cell.value for cell in peer[1]]
    relation_column = peer_headers.index("Relation") + 1
    assert peer.cell(2, relation_column).value == "UNRESOLVED"
    assert peer.conditional_formatting

    breakdowns = workbook["Revenue_Breakdowns"]
    assert breakdowns.cell(2, 7).value == "NOT_COMPARABLE"
    assert breakdowns.cell(2, 10).hyperlink.target.endswith("/amd/amd.htm")

    disclosures = workbook["Disclosure_Status"]
    assert disclosures.cell(2, 3).value == "NOT_REPORTED_THIS_QUARTER"


def test_committed_p2_p3_review_summaries_export_the_full_pilot_without_network(tmp_path: Path) -> None:
    root = Path(__file__).parents[2]
    output = tmp_path / "AMD_MSFT_META_pilot.xlsx"
    export_documented_review_workbook(
        manifest=root / "docs/pilots/amd-msft-meta-filing-manifest.json",
        p2_summary=root / "docs/pilots/amd-msft-meta-p2-dossiers.md",
        p3_summary=root / "docs/pilots/amd-msft-meta-p3-peer-review.md",
        output=output,
    )

    workbook = load_workbook(output)
    assert workbook["Peer_Comparison"].max_row == 16  # header + 15 documented P3 rows
    assert workbook["Disclosure_Status"].max_row == 39  # header + 38 documented P2 states
    assert workbook["Source_Trace"]["S2"].hyperlink.target.endswith("amd-20251227.htm")
